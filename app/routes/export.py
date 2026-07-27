import csv
import io
from datetime import datetime, timedelta

from flask import Blueprint, Response, flash, redirect, url_for

from app.models.employee import Employee
from app.models.schedule import Schedule
from app.models.task import Task
from app.routes.auth import login_required

export_bp = Blueprint('export', __name__)

# --- 1. XUẤT CSV (CŨ) ---
@export_bp.route('/export/tasks')
@login_required
def export_tasks():
    output = io.StringIO()
    output.write('\ufeff')
    writer = csv.writer(output)
    writer.writerow(['Mã nhân viên', 'Họ tên', 'Phòng ban', 'Ngày', 'Ca', 'Công việc', 'Mức độ', 'Thời lượng'])

    schedules = Schedule.query.all()
    for schedule in schedules:
        employee = Employee.query.get(schedule.employee_id)
        task = Task.query.get(schedule.task_id)
        writer.writerow([
            employee.code if employee else '',
            employee.fullname if employee else '',
            employee.department if employee else '',
            schedule.date.strftime('%d/%m/%Y') if schedule.date else '',
            schedule.shift,
            task.task_name if task else '',
            task.priority if task else '',
            task.duration if task else '',
        ])

    response = Response(output.getvalue(), mimetype='text/csv; charset=utf-8')
    response.headers['Content-Disposition'] = 'attachment; filename=task_schedule.csv'
    return response

# --- 2. XUẤT EXCEL THEO FORM MẪU (TÍNH NĂNG MỚI) ---
@export_bp.route('/export/weekly')
@login_required
def export_weekly():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return "Lỗi: Thư viện openpyxl chưa được cài. Vui lòng vào Bash console chạy: pip install openpyxl", 500

    # Lấy ngày của tuần hiện tại
    today = datetime.now().date()
    start_of_week = today - timedelta(days=today.weekday())
    week_dates = [start_of_week + timedelta(days=i) for i in range(7)]
    day_names = ['Thứ Hai', 'Thứ Ba', 'Thứ Tư', 'Thứ Năm', 'Thứ Sáu', 'Thứ Bảy', 'Chủ Nhật']

    wb = Workbook()
    
    # ---------------- SHEET 1: PHÂN CÔNG TUẦN ----------------
    ws1 = wb.active
    ws1.title = "Phân Công Tuần"

    # Định dạng Style Excel
    font_title = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    font_header = Font(name="Arial", size=11, bold=True, color="000000")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    fill_title = PatternFill(start_color="0070C0", fill_type="solid")
    fill_header = PatternFill(start_color="D9E1F2", fill_type="solid")
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Tiêu đề Bảng
    ws1.merge_cells('A1:D1')
    title_cell = ws1['A1']
    title_cell.value = f"BẢNG PHÂN CÔNG CÔNG VIỆC TRONG TUẦN ({week_dates[0].strftime('%d/%m')} - {week_dates[-1].strftime('%d/%m/%Y')})"
    title_cell.font = font_title
    title_cell.alignment = align_center
    title_cell.fill = fill_title
    ws1.row_dimensions[1].height = 30

    # Tiêu đề Cột
    headers = ["Thứ", "Buổi", "Công việc phân công", "Người thực hiện"]
    for col_num, header in enumerate(headers, 1):
        cell = ws1.cell(row=2, column=col_num, value=header)
        cell.font = font_header
        cell.alignment = align_center
        cell.fill = fill_header
        cell.border = border_thin

    # Chỉnh độ rộng cột
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 40
    ws1.column_dimensions['D'].width = 35

    # Lọc dữ liệu tuần từ CSDL
    schedules = Schedule.query.filter(Schedule.date >= week_dates[0], Schedule.date <= week_dates[-1]).all()
    
    current_row = 3
    for i, d in enumerate(week_dates):
        day_str = f"{day_names[i]}\n{d.strftime('%d/%m')}"
        start_row_day = current_row
        has_any_shift = False
        
        for shift in ['Sáng', 'Chiều', 'Tối']:
            tasks_list = [s for s in schedules if s.date == d and s.shift == shift]
            if not tasks_list:
                continue
                
            has_any_shift = True
            start_row_shift = current_row
            
            # Ghi từng công việc ra các dòng
            for s in tasks_list:
                task = Task.query.get(s.task_id) if s.task_id else None
                emp = Employee.query.get(s.employee_id) if s.employee_id else None
                
                t_name = getattr(task, 'task_name', None) or getattr(task, 'name', None) or "Công việc khác"
                e_name = emp.fullname if emp else "Không rõ"
                
                c_task = ws1.cell(row=current_row, column=3, value=t_name)
                c_emp = ws1.cell(row=current_row, column=4, value=e_name)
                
                for c in [c_task, c_emp]:
                    c.border = border_thin
                    c.alignment = align_left
                current_row += 1
                
            # Xử lý Gộp ô (Merge cells) cho cột Buổi
            if current_row - 1 > start_row_shift:
                ws1.merge_cells(start_row=start_row_shift, start_column=2, end_row=current_row-1, end_column=2)
            c_shift = ws1.cell(row=start_row_shift, column=2, value=shift)
            c_shift.alignment = align_center
            for r in range(start_row_shift, current_row):
                ws1.cell(row=r, column=2).border = border_thin

        # Xử lý Gộp ô (Merge cells) cho cột Thứ
        if has_any_shift:
            if current_row - 1 > start_row_day:
                ws1.merge_cells(start_row=start_row_day, start_column=1, end_row=current_row-1, end_column=1)
            c_day = ws1.cell(row=start_row_day, column=1, value=day_str)
            c_day.alignment = align_center
            c_day.font = Font(bold=True)
            for r in range(start_row_day, current_row):
                ws1.cell(row=r, column=1).border = border_thin
        else:
            # Nếu ngày đó trống không có ai làm
            ws1.cell(row=current_row, column=1, value=day_str).alignment = align_center
            ws1.cell(row=current_row, column=1).font = Font(bold=True)
            ws1.cell(row=current_row, column=2, value="Trống").alignment = align_center
            ws1.cell(row=current_row, column=3, value="Chưa có phân công")
            ws1.cell(row=current_row, column=4, value="-").alignment = align_center
            for c in range(1, 5):
                ws1.cell(row=current_row, column=c).border = border_thin
            current_row += 1

    # ---------------- SHEET 2: DANH MỤC CÔNG VIỆC ----------------
    ws2 = wb.create_sheet(title="Danh Mục Công Việc")
    ws2.merge_cells('A1:B1')
    t2 = ws2['A1']
    t2.value = "DANH SÁCH ĐẦU VIỆC HỆ THỐNG"
    t2.font = font_title
    t2.alignment = align_center
    t2.fill = fill_title
    ws2.row_dimensions[1].height = 25

    ws2.cell(row=2, column=1, value="ID").font = font_header
    ws2.cell(row=2, column=2, value="Tên công việc").font = font_header
    for col in [1, 2]:
        ws2.cell(row=2, column=col).fill = fill_header
        ws2.cell(row=2, column=col).border = border_thin
    ws2.cell(row=2, column=1).alignment = align_center

    ws2.column_dimensions['A'].width = 10
    ws2.column_dimensions['B'].width = 50

    tasks = Task.query.all()
    for r, t in enumerate(tasks, 3):
        t_name = getattr(t, 'task_name', None) or getattr(t, 'name', None) or ""
        c1 = ws2.cell(row=r, column=1, value=t.id)
        c2 = ws2.cell(row=r, column=2, value=t_name)
        c1.border = border_thin
        c2.border = border_thin
        c1.alignment = align_center

    # Chuẩn bị file tải xuống
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Đặt tên file thông minh theo Ngày
    file_name = f"PhanCongTuan_{week_dates[0].strftime('%d%m')}_{week_dates[-1].strftime('%d%m%Y')}.xlsx"
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": f"attachment; filename={file_name}"}
    )
